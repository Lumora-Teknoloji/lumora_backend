"""Export all products + latest metrics from database to Excel file."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from datetime import datetime

load_dotenv()

# Database connection
host = os.getenv("POSTGRESQL_HOST", "localhost")
port = os.getenv("POSTGRESQL_PORT", "5432")
db = os.getenv("POSTGRESQL_DATABASE", "bediralvesil_db")
user = os.getenv("POSTGRESQL_USERNAME", "postgres")
pw = os.getenv("POSTGRESQL_PASSWORD", "postgres123")

engine = create_engine(f"postgresql://{user}:{pw}@{host}:{port}/{db}")

print("📊 Veritabanından ürünler çekiliyor...")

with engine.connect() as conn:
    result = conn.execute(text("""
        SELECT 
            p.product_code,
            p.name,
            p.brand,
            p.seller,
            p.url,
            p.image_url,
            p.category,
            p.category_tag,
            p.last_price,
            p.last_discount_rate,
            p.last_engagement_score,
            p.avg_sales_velocity,
            p.review_summary,
            p.sizes,
            p.attributes,
            p.first_seen_at,
            p.last_scraped_at,
            dm.avg_rating,
            dm.rating_count,
            dm.favorite_count,
            dm.cart_count,
            dm.view_count,
            dm.qa_count
        FROM products p
        LEFT JOIN LATERAL (
            SELECT avg_rating, rating_count, favorite_count, cart_count, view_count, qa_count
            FROM daily_metrics dm2
            WHERE dm2.product_id = p.id
            ORDER BY dm2.recorded_at DESC
            LIMIT 1
        ) dm ON true
        ORDER BY p.last_scraped_at DESC
    """))
    rows = result.fetchall()

print(f"✅ {len(rows)} ürün bulundu")

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Tüm Ürünler"

# Headers
headers = [
    "Ürün Kodu", "Ürün Adı", "Marka", "Satıcı", "URL", "Görsel URL",
    "Kategori", "Kategori Etiket", "Fiyat (₺)", "İndirim (%)", 
    "Engagement Score", "Satış Hızı", "Yorum Özeti", "Bedenler", 
    "Özellikler", "İlk Görülme", "Son Güncelleme",
    "Puan", "Yorum Sayısı", "Favori", "Sepet", "Görüntülenme", "S&C Sayısı"
]

# Style headers
header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data rows
for row_idx, row in enumerate(rows, 2):
    for col_idx, value in enumerate(row, 1):
        if isinstance(value, dict):
            value = ", ".join(f"{k}: {v}" for k, v in value.items() if v)
        elif isinstance(value, list):
            value = ", ".join(str(v) for v in value)
        elif isinstance(value, datetime):
            value = value.strftime("%Y-%m-%d %H:%M")
        
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=10)

# Auto-width columns
for col_idx, header in enumerate(headers, 1):
    max_len = len(header) + 2
    for row_idx in range(2, min(len(rows) + 2, 100)):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(rows) + 1}"

# Save
output_path = os.path.join(os.path.dirname(__file__), "..", "lumora_urunler_rapor.xlsx")
output_path = os.path.abspath(output_path)
wb.save(output_path)
print(f"📁 Excel dosyası kaydedildi: {output_path}")
print(f"📊 Toplam: {len(rows)} ürün, {len(headers)} sütun")
